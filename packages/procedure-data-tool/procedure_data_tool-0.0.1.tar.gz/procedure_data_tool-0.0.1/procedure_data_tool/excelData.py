globals().clear()
from openpyxl import load_workbook
from classes.valve import Valve
from classes.valve2 import Valve2
from classes.valve3 import Valve3
from classes.split import Split
from classes.pump import Pump
from classes.tankreturn import TankReturn
from classes.pit import Pit

def ImportComponents(filepath = '//hanford/data/sitedata/WasteTransferEng/Waste Transfer Engineering/1 Transfers/1C - Procedure Review Tools/MasterProcedureData.xlsx'):
    try:
        wb = load_workbook(filename = filepath, data_only=True)
    except FileNotFoundError as e:
        raise e
    pits = {}
    for row in wb["Pits"].iter_rows(min_row=3, values_only= True):
        pit = row[1]
        nace = row[2]
        nacePMID = row[3]
        label = row[4]
        drain = row[5]
        drainSealPos = row[6]
        pits[pit] = Pit(pit, nace, nacePMID, label, drain, drainSealPos)

    for row in wb["Heaters"].iter_rows(min_row=3, values_only= True):
        heater = row[1]
        pit_name = row[2]
        if heater:
            pits[pit_name].heaters.append(heater)

    for row in wb["Encasement Drain Valves"].iter_rows(min_row=3, values_only= True):
        pit_name = row[1]
        valve = row[2]
        position = row[3]
        pits[pit_name].encasementDrainValve.append(valve)
        pits[pit_name].edvPos.append(position)

    for row in wb["TFSPS PMIDs"].iter_rows(min_row=3, values_only= True):
        pit_name = row[3]
        tfsps = row[1]
        tfsps_pmid = row[2]
        pits[pit_name].tfsps.append(tfsps)
        pits[pit_name].tfsps_pmid.append(tfsps_pmid)  

    component_types = {
            "2-Way-Valve": Valve2,
            "3-Way-Valve": Valve3,
            "Split Point": Split,
            "Pump": Pump,
            "": Valve,
            "Tank Return": TankReturn,
            None: Valve 
    }

    #key = EIN, value = component object
    inventory = {}

    cnx = wb['Connections']
    conections_matrix=cnx["D3:F200"]

    for row in cnx.iter_rows(min_row=3, values_only= True):
        component = row[1]
        component_type = row[2]
        inventory[component] = component_types[component_type](component)
        inventory[component].setPit(row[6])
        inventory[component].setJumper(row[7])

    for component, connections in zip(inventory.values(), conections_matrix):
        for connection in connections:
            if connection.value in inventory:
                if connection.value:
                    component.connect(inventory[connection.value])

    return inventory, pits


        
                    