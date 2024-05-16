import openpyxl
import pandas as pd
import re


class XlsxExtractor:
    def __init__(self, source_xlsx, dest_html_file, show_formulas=False, grid_labels=False, cell_titles=False):
        self.source_xlsx = source_xlsx
        self.dest_html_file = dest_html_file
        self.show_formulas = show_formulas
        self.grid_labels = grid_labels
        self.cell_titles = cell_titles

    def extract(self):
        data_and_formulas = self.extract_data_and_formulas(self.source_xlsx)
        self.save_to_html(data_and_formulas, self.dest_html_file)
        print(f"HTML file successfully saved: {self.dest_html_file}")

    def extract_data_and_formulas(self, file_path):
        workbook = openpyxl.load_workbook(file_path, data_only=False)
        sheet_name = workbook.sheetnames[0]
        sheet = workbook[sheet_name]
        max_row = sheet.max_row
        max_column = sheet.max_column

        data = []
        formulas = []

        for row in sheet.iter_rows(max_row=max_row):
            data_row = []
            formula_row = []

            for cell in row:
                data_row.append(cell.value if cell.value is not None else '')
                formula = cell._value if cell.data_type == 'f' else cell.value
                formula_row.append(formula if formula is not None else '')

            data.append(data_row)
            formulas.append(formula_row)

        while data and not any(data[-1]):
            data.pop()
            formulas.pop()

        max_non_empty_cols = max(
            len(row) - next((i for i, val in enumerate(reversed(row))
                             if val != ''), len(row))
            for row in data
        )

        data = [row[:max_non_empty_cols] for row in data]
        formulas = [row[:max_non_empty_cols] for row in formulas]

        # if self.grid_labels then add excel_colname otherwise empty string
        columns = ['' for i in range(max_non_empty_cols)]
        if self.grid_labels:
            columns = [self.excel_colname(i)
                       for i in range(max_non_empty_cols)]

        return {
            'data': pd.DataFrame(data, columns=columns),
            'formulas': pd.DataFrame(formulas, columns=columns)
        }

    def excel_colname(self, n):
        name = ''
        while n >= 0:
            name = chr(n % 26 + ord('A')) + name
            n = n // 26 - 1
        return name

    def compress_html(self, html):
        html = re.sub(r'>\s+<', '><', html)
        html = re.sub(r'\s+', ' ', html)
        html = re.sub(r'<!--.*?-->', '', html)
        return html

    def add_title_attributes(self, df):
        def create_title_attr(val, row_idx, col_idx):
            excel_col = self.excel_colname(col_idx)
            return f'<td title="{excel_col}{row_idx}">{val}</td>'

        rows = df.values.tolist()
        columns = df.columns.tolist()

        new_html = '<table border="1" class="dataframe">\n  <thead>\n    <tr style="text-align: right;">\n'
        new_html += ''.join([f'      <th>{col}</th>\n' for col in columns])
        new_html += '    </tr>\n  </thead>\n  <tbody>\n'

        for row_idx, row in enumerate(rows):
            new_html += '    <tr>\n'
            for col_idx, val in enumerate(row):
                new_html += create_title_attr(val, row_idx + 1, col_idx)
            new_html += '    </tr>\n'

        new_html += '  </tbody>\n</table>'

        return self.compress_html(new_html)

    def add_grid_labels(self, df):
        rows = df.values.tolist()
        columns = df.columns.tolist()

        new_html = '<table border="1" class="dataframe">\n  <thead>\n    <tr style="text-align: right;">\n'
        new_html += '<th></th>' + \
            ''.join([f'<th>{col}</th>\n' for col in columns])
        new_html += '    </tr>\n  </thead>\n  <tbody>\n'

        for row_idx, row in enumerate(rows):
            new_html += f'    <tr><td>{row_idx + 1}</td>\n'
            for col_idx, val in enumerate(row):
                new_html += f'      <td>{val}</td>\n'
            new_html += '    </tr>\n'

        new_html += '  </tbody>\n</table>'

        return self.compress_html(new_html)

    def save_to_html(self, data_dict, output_file):
        df = data_dict['data']

        if self.grid_labels:
            data_html = self.add_grid_labels(df)

            if self.cell_titles:
                data_html = self.add_title_attributes(df)
        else:
            data_html = df.to_html(index=self.grid_labels, border=0)

        html = data_html

        if self.show_formulas:
            df = data_dict['formulas']
            if self.cell_titles:
                formulas_html = self.add_title_attributes(df)
            else:
                formulas_html = df.to_html(index=self.grid_labels, border=0)
            html += "<h1>Formulas</h1>" + formulas_html

        with open(output_file, "w", encoding="utf-8") as data_file:
            data_file.write("<h1>Sheet 1 - Data</h1>")
            data_file.write(html)


if __name__ == "__main__":
    cls = XlsxExtractor(
        source_xlsx='data.xlsx',
        dest_html_file='result.html',
        show_formulas=False,
        grid_labels=True,
        cell_titles=True
    )
    cls.extract()
