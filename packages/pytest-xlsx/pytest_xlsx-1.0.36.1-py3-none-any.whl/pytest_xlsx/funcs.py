import json
import logging
from collections import OrderedDict, defaultdict
from itertools import chain
from pathlib import Path
from typing import Iterable, Optional, Dict, List, AnyStr

import yaml
from openpyxl import load_workbook
from openpyxl.worksheet._read_only import ReadOnlyWorksheet

from pytest_xlsx.settings import settings

logger = logging.getLogger(__name__)


def remove_none_by_dict(
    d: Dict[Optional[AnyStr], List[AnyStr]],
) -> Dict[AnyStr, List[AnyStr]]:
    """移除value中的None，同时移除空value的key"""
    new_d = {}
    for k, v_list in d.items():
        new_v_list = [_ for _ in v_list if _ is not None]
        if not new_v_list:
            continue
        else:
            new_d[k] = new_v_list

    return new_d


def clean_dict(d: Dict[Optional[AnyStr], List[AnyStr]]) -> Dict[AnyStr, str]:
    """如果value是只有一个item的列表，则将修value改为item"""
    new_d = {}

    for k, v in d.items():
        if k is None:
            new_d["_BlankField"] = v  # fix: 不使用None作为字典的key
        elif len(v) == 1:
            new_d[k] = v[0]
        else:
            new_d[k] = v

    return new_d


def iter_xlsx_sheet(path: Path) -> Iterable[ReadOnlyWorksheet]:
    wb = load_workbook(path, read_only=True)
    for ws in wb.worksheets:
        if ws.sheet_state == "hidden":
            logger.debug(f"{ws.title} state  is hidden, so skip")
            continue
        else:
            yield ws


def get_value_by_sequence(*args):
    """

    [["username, id"], "password"] ->  ["username, id", "password"]
    [[1,9],[2,99],[3,999]]    -> [1,9,2,99,3,999]

    :param args:
    :return:
    """
    for arg in args:
        if type(arg) in [list, tuple]:
            yield from arg
        else:
            yield arg


class SheetHandler:
    meta_list = [
        "name",
        "mark",
    ]

    def __init__(self, ws: ReadOnlyWorksheet, debug=False):
        logger.debug(f"SheetHandler: {ws.parent._archive.fp.name}  -> {ws.title}")
        self.ws = ws
        self.meta_col_name = settings.meta_column_name
        self.case_list = []
        self.last_content_type = ""
        self.current_case_no = self.current_step_no = 0
        self.current_case = {
            "id": self.current_case_no,
            "meta": {},
            "steps": [],
        }

        self.headers_map = self._get_headers_map()

        if debug is False:
            self.contents = list(self._get_contents())
            self.make_case_list(self.contents)

    def _get_headers_map(self) -> Dict[Optional[AnyStr], List[int]]:
        logger.debug("加载表头")
        for row_index, row_content in enumerate(
            self.ws.iter_rows(values_only=True, max_row=1),
        ):
            _headers = defaultdict(list)
            for index, col in enumerate(row_content):
                _headers[col].append(index)

            headers_map = dict(_headers)
            logger.debug(f"row_index={row_index}, headers_map={headers_map}")

            if self.meta_col_name not in headers_map:
                msg = f"meta_column_name 不存在，meta_column_name={self.meta_col_name}, headers_map={headers_map}"
                logger.error(msg)
                raise ValueError(msg)

            return dict(_headers)

    def _get_contents(self) -> List[Dict[Optional[AnyStr], List[AnyStr]]]:
        logger.debug("加载表内容")
        for row_index, row_content in enumerate(
            self.ws.iter_rows(values_only=True, min_row=2),
            start=2,
        ):
            if set(row_content) < {None, ""}:
                # 20231228 忽略无内容的行
                logger.debug(f"row_index={row_index}, content ={{None, ''}}")
                continue
            _content = defaultdict(list)
            for k, v in self.headers_map.items():
                for i in v:
                    _content[k].append(row_content[i])
            content = OrderedDict(_content)
            logger.debug(f"row_index={row_index}, content={content}")
            yield content

    def _parse_content(self, content: Dict[Optional[AnyStr], List[AnyStr]]):
        ...

        content_type: str

        meta = content.get(self.meta_col_name, ["NOT_META"])[0]

        if meta in self.meta_list:
            content_type = "meta"
        else:
            content_type = "step"

        if content_type == "meta":
            content = remove_none_by_dict(content)
            if self.last_content_type != content_type:
                # 新的用例
                logger.debug("识别到新的用例")
                self.current_case_no += 1
                self.current_step_no = 0
                self.current_case = {
                    "id": self.current_case_no,
                    "meta": {},
                    "steps": [],
                }
                self.case_list.append(self.current_case)

            meta_key = meta
            meta_index = list(content.keys()).index(self.meta_col_name)
            meta_value = list(chain.from_iterable(content.values()))[meta_index + 1 :]

            self.current_case["meta"].setdefault(meta_key, [])
            self.current_case["meta"][meta_key].append(meta_value)

            logger.debug(f"add meta: {meta_key}.{meta_value}")

        elif content_type == "step":
            step_value = clean_dict(content)  # 20230228，用户自动补全表头
            self.current_case["steps"].append(step_value)
            logger.debug(f"add step: {step_value}")

        self.last_content_type = content_type
        return content_type

    def make_case_list(self, contents):
        for content in contents:
            self._parse_content(content)

    def to_dict(self):
        return OrderedDict({"name": self.ws.title, "case_list": self.case_list})


class LoadData:
    """
    尝试多种方式加载数据：
        - json加载
        - 字符串分割
        - json文件加载
        - yaml文件加载
    """

    row_data: str
    _data: [] = None

    def __init__(self, path: Path, row_data: str):
        self.path = path
        self.row_data = row_data

    @property
    def data(self):
        if self._data is None:
            for f in (self.try_json, self.try_str, self.try_file):
                try:
                    d = f()
                    if isinstance(d, list):
                        self._data = d
                        break
                except Exception:
                    pass

            else:
                raise ValueError(f"参数解析失败: {self.row_data}")
        logger.debug(f"参数解析成功: {self._data}")
        return self._data

    def try_json(self):
        return json.loads(self.row_data)

    def try_str(self):
        if "," in self.row_data:
            return self.row_data.replace(" ", "").split(",")

    def try_file(self):
        # todo 支持绝对路径
        path = self.path.parent / self.row_data
        if not path.exists():
            return

        data = yaml.safe_load(path.read_text(encoding="utf-8"))
        return data
