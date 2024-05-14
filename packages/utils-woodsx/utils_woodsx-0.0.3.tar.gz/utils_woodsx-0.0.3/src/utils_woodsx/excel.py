from openpyxl import load_workbook

def get_excel_data(file_path: str, sheet_name: str = None):
    """
    Get the data from excel file.
    @param file_path: str, the path of excel file.
    @param sheet_name: str, the name of sheet.
    @return: list, the data of excel file.
    """

    wb = load_workbook(file_path)
    
    sheet = wb.active

    if sheet_name is not None:
        sheet = wb[sheet_name]

    data = []

    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    
    return data

def excel2mdtablestr(excel_file_path: str):
    """
    Convert excel file to markdown String with table format.
    @param excel_file_path: str, the path of excel file.
    @return: str, the markdown string with table format.
    """
    
    data = get_excel_data(excel_file_path)

    if data is None or len(data) == 0:
        return None
    
    # get the 1st row as header
    header = data[0]
    header_str = "|" + "|".join([str(h) for h in header]) + "|" + "\n"

    # get the other rows as content
    content = data[1:]
    for i in range(len(content)):
        content[i] = "|" + "|".join([str(c) for c in content[i]]) + "|"

    content_str = "\n".join(content)

    return header_str + "|".join(["----" for g in header]) + "\n" + content_str

excel2mardowntablestr = excel2mdtablestr