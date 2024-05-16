import openpyxl
import os


class Xls_File:
    def __init__(self, file_path):
        self.file_path = file_path

    def xls_open(self):
        # global wb
        if os.path.exists(self.file_path):
            self.wb = openpyxl.load_workbook(self.file_path)
        else:
            self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        # return goal_sheet

    def xls_write(self, i, j, xxxx):
        self.ws.cell(row=i, column=j, value=xxxx)  # goalworksheet['A1'] = 'Hello, world!'

    def xls_value(self, i, j):  # 获取特定单元格的值
        cell_value = self.ws.cell(row=i, column=j).value
        return cell_value

    def xls_append_row(self):
        """
        Get the max blank row of the whole excel
        :return:
        """
        row_site = self.ws.max_row + 1
        return row_site

    def xls_get_next_empty_row_in_column(self, column_index):
        """
        Get the pointed_column's the minimum blank row
        :param column_index:
        :return:
        """
        max_row_with_value = 0
        for row in self.ws.iter_rows(min_col=column_index, max_col=column_index):
            if row[0].value is not None:
                max_row_with_value = row[0].row
        return max_row_with_value + 1

    def xls_get_next_empty_column_in_row(self, row_index):
        """
        Get the pointed_row's the minimum blank column
        :param row_index:
        :return:
        """
        max_col_with_value = 0
        for cell in self.ws[row_index]:
            if cell.value is not None:
                max_col_with_value = cell.column
        return max_col_with_value + 1

    def xls_append_column(self):
        column_site = self.ws.max_column + 1
        return column_site

    def xls_close(self):
        self.wb.save(self.file_path)
        self.wb.close()
        print(os.path.abspath(self.file_path))

    def xls_save(self):
        self.wb.save(self.file_path)

    def xls_create_sheet(self, pSheetName):
        if pSheetName in self.wb:
            self.sheet = self.wb[pSheetName]
        else:
            self.sheet = self.wb.create_sheet(pSheetName)
        return self.sheet

    def xls_active_sheet(self):
        self.wb.active = self.sheet
        self.ws = self.wb.active

    def write_fix_column(self,column, data):
        """
        write data to fixed column and 1 2 3 4 5 row
        向固定列的最小空白行写入数据
        :param column:
        :param data:
        :return:
        """
        # ins_xls_ = Xls_File(self.file_path)
        # ins_xls_.xls_open()
        row = ins_xls_.xls_get_next_empty_row_in_column(column)
        ins_xls_.xls_write(row, column, data)
        # ins_xls_.xls_close()

    def write_fix_row(self,row, data):
        """
         write data to fixed row and 1 2 3 4 5 column
         向固定行的最小空白列写入数据
        :param row:
        :param data:
        :return:
        """
        # ins_xls_ = Xls_File(self.file_path)
        # ins_xls_.xls_open()
        column = ins_xls_.xls_get_next_empty_column_in_row(row)
        ins_xls_.xls_write(row, column, data)
        # ins_xls_.xls_close()


if __name__ == '__main__':
    ins_xls_ = Xls_File("tset123.xlsx")
    ins_xls_.xls_open()

    ins_xls_.write_fix_column(2, 1)
    ins_xls_.write_fix_column(2, 1)
    ins_xls_.write_fix_column(2, 1)
    ins_xls_.write_fix_column(2, 1)
    ins_xls_.xls_close()