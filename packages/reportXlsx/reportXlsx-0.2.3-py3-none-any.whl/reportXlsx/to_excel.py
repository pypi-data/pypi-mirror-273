# -*- coding:utf-8 -*- #

import pandas as pd
import time
import ast
from datetime import datetime,timedelta,date
from openpyxl import Workbook as workbook,load_workbook
import xlwings
from copy import copy
import os
from pathlib import Path
from redmail import EmailSender
import configparser
import pkg_resources
from pythonnet import set_runtime

#os.environ['PYTHONNET_RUNTIME'] = 'coreclr'
#os.environ['PYTHONNET_CORE_ROOT'] = '/usr/share/dotnet/shared/Microsoft.NETCore.App/6.0.30/'
# 配置 pythonnet 使用 .NET Core/.NET 6.0 运行时
#set_runtime(os.environ['PYTHONNET_CORE_ROOT'])
os.environ['PYTHONNET_PYDLL'] = '/usr/share/dotnet/shared/Microsoft.NETCore.App/6.0.30/System.Private.CoreLib.dll'
os.environ['PYTHONNET_RUNTIME'] = 'coreclr'

DLL_FILES = [pkg_resources.resource_filename('reportXlsx', f'resources/{dll_name}') for dll_name in [
    'Spire.XLS.dll','SkiaSharp.dll'
]]
Dll_CLASS = [pkg_resources.resource_filename('reportXlsx',f'resources/{class_name}') for class_name in ['Spire.XLS','SkiaSharp']]
import clr
clr.FindAssembly(DLL_FILES[0])
clr.FindAssembly(DLL_FILES[1])
clr.AddReference(Dll_CLASS[0])
clr.AddReference(Dll_CLASS[1])
from Spire.Xls import *
from openpyxl.styles import numbers
from System.IO import *
from Spire.Xls.Core.Spreadsheet import HTMLOptions




class sqlResult_to_Excel:

    def __init__(self,figure_chart=None,rank_chart=None,
                 left_on=1,
                 right_on=1,
                 sorted_col=6,
                 send_date= datetime.now()-timedelta(days=1),
                 *args
                 ):
        self.figure_chart = figure_chart #通报前半部分
        self.rank_chart = rank_chart #通报的后半部分
        self.left_on = left_on #左表join的行
        self.right_on = right_on #右表关联的行
        self.sorted_col = sorted_col #排序列
        self.send_date = send_date
    
    def _copy_sheet(self,source_ws, target_ws):
    # 复制数据和样式
        for row in source_ws.iter_rows():
            for cell in row:
                new_cell = target_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = cell.number_format
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)

        # 处理合并单元格
        for merge_cell in source_ws.merged_cells.ranges:
            target_ws.merge_cells(str(merge_cell))
    
    def solve_cannot_auto_calculate_func(self,path):
        excel_app = xlwings.App(visible=False)
        excel_book = excel_app.books.open(path)
        excel_book.save()
        excel_book.close()
        excel_app.quit()
    
    def replace_value(self,value):
        if value == 290:
            return '西安'
        elif value == 910:
            return '咸阳'
        elif value == 911:
            return '延安'
        elif value == 912:
            return '榆林'
        elif value == 913:
            return '渭南'
        elif value == 914:
            return '商洛'
        elif value == 915:
            return '安康'
        elif value == 916:
            return '汉中'
        elif value == 917:
            return '宝鸡'
        elif value == 919:
            return '铜川'
        else:
            return '全省'
    def deal_with_sql_result(self,need_delete_col_list=[0,7]):
        figure_chart = pd.DataFrame(self.figure_chart)
        rank_chart = pd.DataFrame(self.rank_chart)
        merge_chart = pd.merge(figure_chart,rank_chart,left_on=self.left_on,right_on=self.right_on,how='left').reindex()
        #针对merge_chart排序
        part1 = merge_chart.iloc[:-1].sort_values(by= self.sorted_col,ascending= False)
        last_row = merge_chart.iloc[-1:]
        df_result = pd.concat([part1,last_row],ignore_index=True)
        df_result.columns = list(range(len(df_result.columns)))
        #替换文字
        df_result[1]=df_result[1].apply(self.replace_value)
        df_result.drop(need_delete_col_list,axis= 1 ,inplace = True)
        df_result = df_result.reset_index(drop = True)
        return df_result
    def to_excel(self,
                df_result,
                 excl_start_row = 4,
                 excl_start_col =1,
                 excl_file_path='',
                 **pos):
        wb = load_workbook(excl_file_path)
        sheet_name = self.sheet_name
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title= self.sheet_name)

        start_row = excl_start_row
        start_col = excl_start_col

        for r_idx,row in df_result.iterrows():
            for c_idx,value in enumerate(row):
                ws.cell(row=start_row+r_idx,column=start_col+c_idx,value = value)
        
        wb.save(excl_file_path)

    def update_by_template(self,codilist,metalist,path='template.xlsx',dis ='',sheet_name = '通报'):
        wb = load_workbook(path)
        sh_name = sheet_name
        #print(sh_name)
        if sh_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title= sheet_name)

        for item1,item2 in zip(codilist,metalist):
            start_row = item1[0]
            start_col = item1[1]
            #print(df_tmp)
            for r_idx,row in enumerate(item2):
                #print(row),print(type(row))
                for c_idx,value in enumerate(ast.literal_eval(row)):
                    #print(value)
                    ws.cell(row=start_row+r_idx,column=start_col+c_idx,value = value)
                    #print(ws.cell(row=start_row+r_idx,column=start_col+c_idx,value = value).value)
                
                    
        if not dis:
            wb.save(path)
        else:
            wb.save(dis)
    
    def ajust_and_copy_excel(self,source_codilist,target_codilist,path='template_new_v1.xlsx',sheet_name = '通报'):
        self.solve_cannot_auto_calculate_func(path)

        wb = load_workbook(path,data_only=True)
        sheet = wb[sheet_name]
        wb.save('test.xlsx')
        #针对所有坐标开展排序工作
        for s_codi,t_codi in zip(source_codilist,target_codilist):
            source_range = sheet[s_codi]
            target = sheet[t_codi]
            start_row = target.row
            start_col = target.column
            for source_row_index,source_row in enumerate(source_range):
                for source_col_index,source_cell in enumerate(source_row):
                    target_row = start_row + source_row_index
                    target_col = start_col + source_col_index
                    #print(target_row,target_col)
                    target_cell = sheet.cell(row=target_row,column=target_col)
                    target_cell.value = source_cell.value
                    
        wb.save('template_new_v2.xlsx')

    def sort_excel(self,sort_range_list,sort_col_index_list,path,if_reverse,sheet_name = '通报'):
        wb = load_workbook(path,data_only=True)
        sheet = wb[wb.sheetnames[0]]
        for s_range ,sort_index in zip(sort_range_list,sort_col_index_list):

            sort_range = sheet[s_range]
            data = list(sort_range)
            data_all=[]
            for index,row in enumerate(data):
                data_row =[]
                for item in enumerate(row):
                    data_row.append(item[1].value)
                data_all.append(data_row)
            data_all.sort(key = lambda row:row[sort_index],reverse=if_reverse)
        
            start_row = sort_range[0][0].row  # 范围的起始行号
            start_col = sort_range[0][0].column  # 范围的起始列号

            for row_index, row_data in enumerate(data_all):
                for col_index,value in enumerate(row_data):
                    # 计算目标单元格的行和列位置
                    target_row = start_row + row_index
                    target_col = start_col + col_index
                    # 将排序后的值写入目标单元格
                    sheet.cell(row=target_row, column=target_col).value = value

        wb.save('template_new_v3.xlsx')
        wb.save('三级今日通报数据.xlsx')

    #def excel_to_pic_index_less_equal_2(self,to_pic_generate_range,path = '',sheet_name = '通报',pic_name = 'results.png'):

        ##这是一个生成通报图片的函数
        ##workbook = Workbook()
        ##workbook.LoadFromFile(path)
        ##sheet= workbook.Worksheets[0]
        ##image = sheet.ToImage(
            ###0,0,86,8
            ##to_pic_generate_range[0],
            ##to_pic_generate_range[1],
            ##to_pic_generate_range[2],
            ##to_pic_generate_range[3]
        ##)
        ##image.Save('cellrangeImage.png',ImageFormat.get_Png())
        ##workbook.Dispose()



        #wb = Workbook()
        #wb.LoadFromFile(path)
        #if wb.Index < 2:
            #sheet = wb.Worksheets[sheet_name]
            #setting = ConverterSetting()
            #setting.XDpi = 300
            #setting.YDpi = 300
            #wb.ConverterSetting =setting
            #sheet.SaveToImage(pic_name,to_pic_generate_range[0],to_pic_generate_range[1],to_pic_generate_range[2],to_pic_generate_range[3] )
        
        #wb.Dispose()

    
   
    def excel_to_html(self,to_pic_generate_range,path = '',sheet_name = '通报'):
        #wb = Workbook()
        #wb.LoadFromFile(path)

        #sheet = wb.Worksheets[sheet_name]
        #sheet.SaveToImage(pic_name,to_pic_generate_range[0],to_pic_generate_range[1],to_pic_generate_range[2],to_pic_generate_range[3] )
        #return html
        wb = Workbook()
        source_wb = Workbook()
        source_wb.LoadFromFile(path)
        old_sheet = source_wb.Worksheets[sheet_name] 
        # 选取表格的范围 
        cell_range = old_sheet.Range[to_pic_generate_range] 
        sheet = wb.Worksheets[0]
        cell_range.Copy(sheet.Range[to_pic_generate_range])
        
        stream = MemoryStream()
        options = HTMLOptions()
        options.SavedAsFragment = True
        options.isExportStyle = True
        sheet.SaveToHtml(stream,options)

        stream.Position = 0;
        reader = StreamReader(stream);
        htmlString = reader.ReadToEnd();
        wb.Dispose()
        return htmlString

    def del_col_excl(self,starting_col_index,need_delete_cols,path='',sheet_name ='通报'):
        
        wb =load_workbook(path)
        sh_name = sheet_name
        if sh_name in wb.sheetnames:
            ws = wb[sh_name]

        starting_cols_index = starting_col_index
        num_of_cols_to_delete = need_delete_cols
        ws.delete_cols(starting_cols_index,num_of_cols_to_delete)

        wb.save(path)


    def add_detail(self,results_df_list,name_lists,path='template_new_v3.xlsx',target = ''):
        if not os.path.exists(path) or not path:
            wb = workbook()
            default_sheet = wb.active
            wb.remove(default_sheet)
        else:
            wb = load_workbook(path)
        for item,name in zip(results_df_list,name_lists):
            new_sheet = wb.create_sheet(title=name)
            for row in item:
                str_row = tuple(str(item) for item in row)
                new_sheet.append(str_row)
        if not target:
            wb.save(path)
        else:
            wb.save(target)

    def set_up_style_of_the_workbook(self,sheet_need_change_list,need_set_col_tuple,path=''):
        wb = load_workbook(path)
        for sheet,need_row in zip(sheet_need_change_list,need_set_col_tuple):
            ws = wb[sheet]  
            # 在单元格 B1 中设置值为 0.54321
            #ws['B1'] = 0.54321
            max_row = ws.max_row
            max_col = ws.max_column

            for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
                for cell in row:
                    # 获取单元格的文本值
                    text_value = cell.value
                    # 如果文本值可以转换为数字，则进行转换并覆盖原始值
                    try:
                        numeric_value = float(text_value)
                        cell.value = numeric_value
                    except (ValueError, TypeError):
                        # 如果无法转换为数字，则保持原始值不变
                        pass
             
            print(need_row)

            for column in need_row:
                print(need_row)
                for row in range(1, max_row + 1):
                    cell = ws.cell(row=row, column=column)
                    cell.number_format = numbers.FORMAT_PERCENTAGE

        # 保存工作簿
        wb.save(path)
        # 设置 B1 单元格为百分比格式
        #ws['I5'].number_format = numbers.FORMAT_NUMBER
        #ws['I5'].number_format = '0.00%'

        ## 保存工作簿
        #wb.save('formatted.xlsx')
    def split_the_excel(self,target_excel,column_index = 1,target_sheet =''):
        wb = load_workbook(target_excel)
        sheet = wb[target_sheet]
        ws = wb.active
        # 创建一个字典，用于存储按照字段值拆分后的数据
        split_data = {}
        #遍历每一行，根据指定字段的值进行拆分
        for row in ws.iter_rows(min_row=2, values_only=True):
            key = row[column_index - 1]  # 由于列索引从1开始，所以需要减1
            if key not in split_data:
                split_data[key] = []
            split_data[key].append(row)

        # 根据拆分后的数据创建新的Excel文件
        for key, data in split_data.items():
            wb_new = Workbook()
            ws_new = wb_new.active
            ws_new.append(ws[1])  # 复制原始Excel的标题行
            for row in data:
                ws_new.append(row)
            wb_new.save(f"{key}.xlsx")

    def add_sheet(self,target_wb='',need_add_workbook='',need_sheet_name='',new_sheet_name='new sheet'):
        #添加sheet
        wb = load_workbook(target_wb)
        wb_old = load_workbook(need_add_workbook)
        ws_old = wb_old[need_sheet_name]
        ws_new = wb.create_sheet(new_sheet_name)

        self._copy_sheet(ws_old,ws_new)
        wb.save(target_wb)
    
    def add_sheet_with_style(self,target_wb='',source_workbook='',source_sheet_name='',new_sheet_name='new sheet'):
        #添加sheet
        wb = Workbook()
        wb.LoadFromFile(target_wb)
        if wb.Worksheets[new_sheet_name]:
            wb.Worksheets.Remove(wb.Worksheets[new_sheet_name])
        source_wb = Workbook()
        source_wb.LoadFromFile(source_workbook)
        old_sheet = source_wb.Worksheets[source_sheet_name] 
        wb.Worksheets.AddCopy(old_sheet)

        wb.Worksheets[source_sheet_name].Name = new_sheet_name
        new_sheet = wb.Worksheets[new_sheet_name]
        wb.ActiveSheetIndex = wb.Worksheets.Count - 1
        wb.SaveToFile('final_excel_v2.xlsx',FileFormat.Version2013)
    
    def excel_to_pic(self,to_pic_generate_range,path = '',sheet_name = '通报',pic_name = 'results.png',dpi = 180):
        #生成新xlsx
        wb = Workbook()
        source_wb = Workbook()
        source_wb.LoadFromFile(path)
        old_sheet = source_wb.Worksheets[sheet_name] 
        wb.Worksheets.AddCopy(old_sheet)
        wb.Worksheets.RemoveAt(0)
        wb.Worksheets.RemoveAt(1)
       
        setting = ConverterSetting()
        setting.XDpi = dpi
        setting.YDpi = dpi
        
        wb.ConverterSetting =setting
        sheet = wb.Worksheets[sheet_name]
        sheet.SaveToImage(pic_name,to_pic_generate_range[0],to_pic_generate_range[1],to_pic_generate_range[2],to_pic_generate_range[3] )
        wb.Dispose()
        



    


if __name__ =="__main__":
    #slq = sqlResult_to_Excel()
    #slq.add_sheet_with_style(target_wb='final_excel_v1.xlsx',source_workbook='template_new_v3.xlsx',source_sheet_name='通报',new_sheet_name='0428')

    ##slq.excel_to_pic((1,1,47,15),path='final_excel_v1.xlsx',sheet_name='0428')
    ##slq.excel_to_pic_index_gt_2((1,1,47,15),path='template_new_v3.xlsx')
    #slq.excel_to_html('A1:K47',path='template_new_v3.xlsx')
    #mail_config = configparser.ConfigParser()

    #mail_config.read('mail_config.ini')
    #email = EmailSender(
        #host= 'smtp.chinatelecom.cn',
        #port= 587,
        #username= mail_config['sender']['username'],
        #password= mail_config['sender']['password'],
        #use_starttls= False

    #)

    #send_mail_client = send_mail(email=email)
    #excel_name =  "239套餐降档流失日通报_" +(date.today()-timedelta(days=1)).strftime("%Y年%m月%d日")+'.xlsx'
    #html_text = slq.excel_to_html('A1:K47',path='template_new_v3.xlsx')
    #send_mail_client.send_mail(
        #sender='wangyinping@chinatelecom.cn',
        ##recv_list=["xazhaoxiaoli@chinatelecom.cn","xabaibo@chinatelecom.cn",
                    ##"xabianying@chinatelecom.cn",
                    ##"xywangdong@chinatelecom.cn",
                    ##"xychenliang@chinatelecom.cn",
                    ##"wnyangjingsen@chinatelecom.cn",
                    ##"wnjiyanfang@chinatelecom.cn",
                    ##"wncaoying@chinatelecom.cn",
                    ##"bjlimei@chinatelecom.cn",
                    ##"bjzengzili@chinatelecom.cn",
                    ##"hzmaxinping@chinatelecom.cn",
                    ##"hzdujianjun@chinatelecom.cn",
                    ##"hzzhangting@chinatelecom.cn",
                    ##"ylheixifeng@chinatelecom.cn",
                    ##"ylyanghuixia@chinatelecom.cn",
                    ##"ylgaolei01@chinatelecom.cn",
                    ##"akpanli@chinatelecom.cn",
                    ##"akyangwenling@chinatelecom.cn",
                    ##"akpurui@chinatelecom.cn",
                    ##"slzhoufang@chinatelecom.cn",
                    ##"sllishasha@chinatelecom.cn",
                    ##"tcxunping@chinatelecom.cn",
                    ##"15399191880@189.cn",
                    ##"yalimei@chinatelecom.cn",
                    ##"yachenna@chinatelecom.cn",
                    ##"sunpp@chinatelecom.com.cn"
    ##],
        #recv_list=['yabailu@chinatelecom.cn','wangyinping@chinatelecom.cn','xawuyingying@chinatelecom.cn'],
        ##cc_list=['dongxiaochang@chinatelecom.cn','huangmin@chinatelecom.cn','lixilin@chinatelecom.cn',
                ##'xyguanyu@chinatelecom.cn',
                ##'yabaibaohong@chinatelecom.cn',
                ##'doule@chinatelecom.cn',
                ##'yabailu@chinatelecom.cn',
                ##'quxiaogang@chinatelecom.cn',
                ##'wangdianwei@chinatelecom.cn',
                ##'wnzuojiao@chinatelecom.cn',
                ##'tongwenhui@chinatelecom.cn',
                ##'fanrl@chinatelecom.cn',
                ##'yaoyc5@chinatelecom.cn',
                ##'guor2@chinatelecom.cn',
                ##'wangyinping@chinatelecom.cn'
             
                ##],
        #cc_list=['xawuyingying@chinatelecom.cn'],
        #subject="239套餐降档流失日通报_" +(date.today()-timedelta(days=1)).strftime("%Y年%m月%d日"),
        ##text=(date.today()-timedelta(days=1)).strftime("%Y年%m月%d日") + "日报"
        #html= html_text,
        ##html='''
            ##<p></p >
            ##{{report}}
        ##''',
        ##body_images={'report':'results.png'}
        #attachments={
            #excel_name:"final_excel_v1.xlsx"
        #}
        #)
    print("日报处理完毕")






        
        
        






    

    



##################-----------test-------------###############
#test_case 
#wb_excl = sqlResult_to_Excel()
#wb_excl.excel_to_pic((0,0,86,9),'your_excel_file.xlsx')




#wb_excl = sqlResult_to_Excel()
##wb_excl.update_by_template(
    ##codi,
    ##results,
    ##path='template_new.xlsx'
##)
##source = ['K4:R14','K18:R28','K47:R57','K61:R71','K74:R84']
##target = ['A4','A18','A47','A61','A74']
## 排序

#source = ['A4:H13','A18:H27','A47:H56','A61:H70','A74:H83']
#sort_col =[5,5,5,5,4]


##wb_excl.ajust_and_copy_excel(source,target,path='template_new.xlsx')
##wb_excl.sort_excel(source,sort_col,'template_new.xlsx',True)
#wb_excl.excel_to_pic((0,0,86,8),path='your_excel_file.xlsx')
#wb_excl.excel_to_pic_v1()







        


        